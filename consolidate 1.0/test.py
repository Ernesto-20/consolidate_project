import pandas as pd
from openpyxl import load_workbook

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

template_fout = "prueba_2"
fout_name = 'DataFiles/New Model/beta_01/Binary Model.xlsb'
save_name = 'DataFiles/New Model/beta_01/Model_2.xlsx'

...
...
...

# df_values is the DataFrame with the data to save
wb = load_workbook(fout_name)
ws = wb['Prod']
print(type(wb))
print(type(ws))
print(ws)


# cell_data = ws.cell(row = 1, column = 1)
# cell_data.value = 'HOLA MUNDO'
#
# wb.save(save_name)


