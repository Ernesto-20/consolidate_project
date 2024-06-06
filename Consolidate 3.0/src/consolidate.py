import os
import datetime
import sys
import time
import tkinter

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from src.manage.resume_div_manage import ResumeDivManage
from src.manage.resume_three_manage import ResumeThreeManage
from src.manage.resume_two_manage import ResumeTwoManage
from src.tables.res1 import Res1
# from src.tools.change_format import ChangeFormat
from src.tools.worker_file import WorkerFile
from src.manage.resume_one_manage import ResumeOneManage
from tkinter import messagebox
# from src.tools.change_format import ChangeFormat


def consolidate(file_list, file_name, file_model, name_model, file_stock):
    # Code for hiding the second windows when show it an error or warning message
    root = tkinter.Tk()
    root.withdraw()
    # End hiding code

    begin = time.time() 

    file_list = file_list
    files_names = file_name
    worker_files_list = []

    if len(file_list) == 0:
        # Exception: no hay elementos en file_list
        messagebox.showwarning(title='Advertencia', message='No se encontraron archivos a consolidar.')
        sys.exit()
        
    file_counter = 0
    date = None
    week_number = None
    all_ids = set() # All Products ids, for example: {'cog': Product}
    ids_cost_dict = {}

    for file in file_list:
        name = files_names[file_counter]
        
        # temp_file_name = ChangeFormat.xlsb_to_xlsx(file)
        # wb = load_workbook(temp_file_name, data_only=True)

        wb = load_workbook(file, data_only=True)
        
        ws_resume_1 = None
        ws_resume_2 = None
        ws_resume_3 = None
        try:
            ws_resume_1 = wb['Res1']
        except KeyError:
            # Exception: No existe la hoja Res1
            messagebox.showerror(title='Error', message=
                                   'El archivo de "'+name+'" no contiene la hoja "Res1".\n\n Puede que en el archivo, el nombre "Res1" '
                                                          'contenga algun espacio en blanco o caracter extraño.\nPor ejemplo: "Res1 "')
            sys.exit()
        try:
            ws_resume_2 = wb['Res2']
        except KeyError:
            # Exception: No existe la hoja Res2
            messagebox.showerror(title='Error', message=
                                   'El archivo de "'+name+'" no contiene la hoja "Res2".\n\n Puede que en el archivo, el nombre "Res2" '
                                                          'contenga algun espacio en blanco o caracter extraño.\nPor ejemplo: "Res2 "')
            sys.exit()
        try:
            ws_resume_3 = wb['Res3']
        except KeyError:
            # Exception: No existe la hoja Res3
            messagebox.showerror(title='Error', message=
                                   'El archivo de "'+name+'" no contiene la hoja "Res3".\n\n Puede que en el archivo, el nombre "Res3" '
                                                          'contenga algun espacio en blanco o caracter extraño.\nPor ejemplo: "Res3 "')
            sys.exit()
        try:
            ws_div = wb['Div']
        except KeyError:
            # Exception: No existe la hoja Div
            messagebox.showerror(title='Error', message=
                                   'El archivo de "'+name+'" no contiene la hoja "Div".\n\n Puede que en el archivo, el nombre "Div" '
                                                          'contenga algun espacio en blanco o caracter extraño.')
            sys.exit()



        is_pivot = name == 'jor'
        worker = WorkerFile(name, ResumeOneManage.get_resume_one(ws_resume_1, is_pivot=is_pivot),
                            ResumeTwoManage.get_resume_two(ws_resume_2, is_pivot=is_pivot),
                            ResumeThreeManage.get_resume_three(ws_resume_3),
                            ResumeDivManage.get_resume_div(ws_div))
        worker_files_list.append(worker)
        
        for key in worker.get_res_1().keys():
            if key not in ids_cost_dict:
                ids_cost_dict[key] = Res1(key, None, None, None, None, None, None, None, None, worker.get_res_1()[key].get_cb(), worker.get_res_1()[key].get_cn())
        all_ids.update(worker.get_res_1().keys())


        if is_pivot:
            date = ws_resume_1.cell(row=2, column=1).value
            week_number = ws_resume_1.cell(row=1, column=3).value

        file_counter += 1
    
    
    del wb
    del ws_resume_1
    # del ws_resumen_2
    # worker_report = to_consolidate_data(worker_files=worker_files_list)
    workbook = None
    try:
        workbook = load_workbook(file_model)
    except FileNotFoundError:
        # Exception: No hay ningun modelo
        messagebox.showerror(title='Error', message=
        'No se encuentra el archivo "Cons-sem__.mgzn.16-05-23.xlsx".\n Si esta el archivo, verifique que el nombre sea correspondiente')
        sys.exit()

    name_counter = 0
    # print('********************')
    # print(len(ids_cost_dict.keys()))
    for worker_file in worker_files_list:
        set_ids = set(worker_file.get_res_1().keys())
        differs = all_ids.difference(set_ids),
        # print('type: differs: ',type(differs)) #differs is a tuple with one set() object
        
        for differ in differs[0]:
            print(differ)
            worker_file.get_res_1()[differ] = Res1(differ, None, None, None, [], [], None, None, None, ids_cost_dict[differ].get_cb(), ids_cost_dict[differ].get_cn())

        # Adding summary information

        # Set Res1 sheet in consolidation
        resume_1_sheet = files_names[name_counter] + '1'
        try:
            resume_1_ws = workbook[resume_1_sheet]
        except KeyError:
            # Exception: No hay una hoja con ese nombre en el modelo.
            messagebox.showerror(title='Error', message=
            'No se encuentra la hoja con nombre "'+ resume_1_sheet+'" en archivo de consolidacion.\n')
            sys.exit()
        is_pivot = worker_file.get_name() == 'jor'
        ResumeOneManage.set_resume_one(resume_1_ws, worker_file, is_pivot=is_pivot)

        # Set Res2 sheet in consolidation
        resume_2_sheet = files_names[name_counter] + '2'
        try:
            resume_2_ws = workbook[resume_2_sheet]
        except KeyError:
            # Exception: No hay una hoja con ese nombre en el modelo.
            messagebox.showerror(title='Error', message=
            'No se encuentra la hoja con nombre "'+ resume_2_sheet+'" en archivo de consolidacion.\n')
            sys.exit()
        ResumeTwoManage.set_resume_one(resume_2_ws, worker_file, is_pivot=is_pivot)

        # Set Res3 sheet in consolidation
        ResumeThreeManage.set_resume_three(resume_2_ws, worker_file)

        # Set Div sheet in consolidation
        ResumeDivManage.set_resume_div(resume_2_ws, worker_file, is_pivot= is_pivot)

        name_counter += 1

    # Adding grost and net cost
    try:
        sem_1_ws = workbook['SEM1']
    except KeyError:
        # Exception: No hay una hoja con ese nombre en el modelo.
        messagebox.showerror(title='Error', message=
        'No se encuentra la hoja con nombre "'+ sem_1_ws+'" en archivo Cons-sem__.mgzn.16-05-23.xlsx.\n')
        sys.exit()
        
    rows_temp = 3
    for product_key in sorted(ids_cost_dict.keys()):
        cb = ids_cost_dict[product_key].get_cb()
        cn = ids_cost_dict[product_key].get_cn()
        sem_1_ws.cell(row=rows_temp, column=50, value=cb)
        sem_1_ws.cell(row=rows_temp, column=51, value=cn)
        rows_temp = rows_temp + 1


    current_Date = datetime.datetime.now()
    date = str(current_Date.day) + '-' + str(current_Date.month) \
           + '-' + str(current_Date.year)[2:] + '  [' + str(current_Date.hour) + \
           ';' + str(current_Date.minute) + ']'

    name_new_report = str(name_model)[:8]+str(week_number)+str(name_model)[10:]+date+'.xlsx'

    print('Guardado: ', name_new_report)
    workbook.save(name_new_report)
    workbook.close()


    # New workbook for stock sheet
    if file_stock != None:
        stock_wb = load_workbook(file_stock)
        try:
            stock_ws = stock_wb['Stock']
            for worker_file in worker_files_list:
                for i in range(1, stock_ws.max_column):
                    if worker_file.get_name() == str(stock_ws.cell( row=2, column=i).value):
                        products = sorted(worker_file.get_res_1().keys())
                        counter_row = 3
                        for key in products:
                            product = worker_file.get_res_1()[key]
                            stock_ws.cell( row=counter_row, column=i, value=product.get_real_existence())
                            counter_row = counter_row + 1
                        break

            products = sorted(worker_file.get_res_1().keys())
            counter_row = 3
            for key in products:
                product = worker_files_list[0].get_res_1()[key]
                stock_ws.cell( row=counter_row, column=1, value=product.get_id())
                counter_row = counter_row + 1
            name_stock = 'Stock-sem' + str(week_number) + '.mgzn.' + date + '.xlsx'
            stock_wb.save(name_stock)
        except KeyError:
            messagebox.showerror(title='Error', message='No se encuentra la hoja con nombre "Stock" en archivo "'+file_stock+'"\n')





    # ChangeFormat.xlsx_to_xlsb(name_new_report)

    # delete all temporal files.
    # os.remove(temp_file_name)
    # os.remove(name_new_report)

    end = time.time()

    print('Tiempo de ejecucion: ', end - begin)
    messagebox.showinfo(title='Listo', message='Todo bien\n')
    sys.exit()

if __name__ == "__main__":
    consolidate()
