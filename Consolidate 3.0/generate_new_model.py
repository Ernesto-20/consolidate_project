import os
import sys
import tkinter
from tkinter import messagebox
from openpyxl import load_workbook

from src.manage.concept_account_manage import ConceptAccountManage
from src.manage.consolidate_resume_manage import ConsolidateResumeManage
from src.manage.info_product.ad_group_manage import ADGroupManage
from src.manage.info_product.bc_group_manage import BCGroupManage
from src.manage.product_price_manage import ProductPriceManage
from src.tables.product_price import ProductPrice
from src.tools.file_model import FileModel
from src.tools.consolidate_tools import load_sheet

# if __name__ == "__main__":

root = tkinter.Tk()
root.withdraw()

consolidate_file = None
pivot_file = None
new_model_file = None

for file in os.listdir('./'):
    if file.__contains__('.xlsx') and file[0] != '~':
        if file.__contains__('Cons-sem'):
            consolidate_file = file

        elif file.__contains__('Cont-aaa'):
            pivot_file = file
        elif file.__contains__('Cont-nom'):
            new_model_file = file

if consolidate_file is None:
    messagebox.showerror(title='Error', message='No se encontró el archivo excel con nombre "Cons-sem..."')
    sys.exit()
if pivot_file is None:
    messagebox.showerror(title='Error', message='No se encontró el archivo excel con nombre "Cont-aaa..."')
    sys.exit()
if new_model_file is None:
    messagebox.showerror(title='Error', message='No se encontró el archivo excel con nombre "Cont-nom..."')
    sys.exit()

consolidate_wb = load_workbook(consolidate_file, data_only=True)
consolidate_ws_resume = load_sheet(consolidate_wb, 'res', consolidate_file)

pivot_wb = load_workbook(pivot_file, data_only=True)
pivot_ws_prod = load_sheet(pivot_wb, 'Prod', pivot_file)
pivot_ws_prec = load_sheet(pivot_wb, 'Prec', pivot_file)
pivot_ws_cue_con = load_sheet(pivot_wb, 'Cue-Con', pivot_file)

new_model_wb = load_workbook(new_model_file)
new_model_ws_prod = load_sheet(new_model_wb, 'Prod', new_model_file)
new_model_ws_prec = load_sheet(new_model_wb, 'Prec', new_model_file)
new_model_ws_cue_con = load_sheet(new_model_wb, 'Cue-Con', new_model_file)

# consolidate_wb = load_workbook(consolidate_file, data_only=True)
# consolidate_ws_res = load_sheet(consolidate_wb, 'res', consolidate_file)
file_model = FileModel(ADGroupManage.get_ad_group(pivot_ws_prod),
                       BCGroupManage.get_bc_group(pivot_ws_prod),
                       ProductPriceManage.get_products_price(pivot_ws_prec),
                       ConceptAccountManage.get_concept_account(pivot_ws_cue_con),
                       ConsolidateResumeManage.get_gross_cost(consolidate_ws_resume))

# Passing Data
ADGroupManage.set_ad_group(new_model_ws_prod, file_model)
BCGroupManage.set_bc_group(new_model_ws_prod, file_model)
ProductPriceManage.set_products_price(new_model_ws_prec, file_model)
ConceptAccountManage.set_concept_account(new_model_ws_cue_con, file_model)
ConsolidateResumeManage.set_gross_cost(new_model_ws_prod, file_model)




new_model_wb.save('saved.xlsx')
print('finish')