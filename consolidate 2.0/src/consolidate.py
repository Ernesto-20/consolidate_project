import os
import datetime
import sys
import time
import tkinter

from openpyxl import load_workbook
from src.manage.balance_manage import BalanceManage
from src.manage.coin_inventory_manage import CoinInventoryManage
from src.manage.coin_control_manage import CoinControlManage
from src.manage.merchandise_control_manage import MerchandiseControlManage
# from src.tools.change_format import ChangeFormat
from src.tools.worker_file import WorkerFile
from src.manage.products_manage import ProductsManage
from tkinter import messagebox


def consolidate(file_list, file_name, file_model, name_model):
    # ui_initialize()
    # Code for hiding the second windows when show it an error or warning message
    root = tkinter.Tk()
    root.withdraw()
    # End hiding code

    begin = time.time()

    file_list = file_list
    files_names = file_name
    worker_files_list = []

    # for file in os.listdir('./'):
    #     if file.__contains__('.xlsb') and file[0] != '~' and file.__contains__('Reporte') == False:
    #         init = file.find('-') + 1
    #         end = file.find('-') + 4
    #         files_names.append(file[init: end])
    #         # files_names.append(file)
    #         file_list.append(file)
    if len(file_list) == 0:
        # Exception: no hay elementos en file_list
        messagebox.showwarning(title='Advertencia', message='No se encontraron archivos a consolidar.')
        sys.exit()
    file_counter = 0
    date = None
    week_number = None
    temp_file_name = None

    for file in file_list:
        # print(file)
        name = files_names[file_counter]
        # temp_file_name = ChangeFormat.xlsb_to_xlsx(file)

        # wb = load_workbook(temp_file_name, data_only=True)
        # wb = load_workbook(file, data_only=True)
        wb = load_workbook(file, data_only=True)
        ws_resumen = None
        try:
            ws_resumen = wb['Resumen']
        except KeyError:
            # Exception: No existe esa hoja
            messagebox.showerror(title='Error', message=
                                   'El archivo de "'+name+'" no contiene la hoja "Resumen".\n\n Puede que en el archivo, el nombre "Resumen no '
                                                          'contenga algun espacio en blanco o caracter extraño.' )
            sys.exit()
        try:
            ws_balance = wb['Balance']
        except KeyError:
            # Exception: No existe esa hoja
            messagebox.showerror(title='Error', message=
                                   'El archivo de "'+name+'" no contiene la hoja "Balance".\n\n Puede que en el archivo, el nombre "Balance no '
                                                          'contenga algun espacio en blanco o caracter extraño.' )
            sys.exit()

        is_pivot = name == 'aaa'
        # if is_pivot: print(file)
        worker_files_list.append(WorkerFile(name, ProductsManage.get_products(ws_resumen, is_pivot=is_pivot),
                                            CoinInventoryManage.get_coin_inventory_source(ws_resumen, is_pivot=is_pivot),
                                            MerchandiseControlManage.get_merchandise_control_source(ws_resumen, is_pivot=is_pivot),
                                            BalanceManage.get_balance_source(ws_balance, is_pivot=is_pivot)
                                            ))
        if is_pivot:
            date = ws_resumen.cell(row=2, column=1).value
            week_number = ws_resumen.cell(row=3, column=1).value

        file_counter += 1
    del wb
    del ws_resumen
    del ws_balance
    # worker_report = to_consolidate_data(worker_files=worker_files_list)
    workbook = None
    try:
        workbook = load_workbook(file_model)
    except FileNotFoundError:
        # Exception: No hay ningun modelo
        messagebox.showerror(title='Error', message=
        'No se encuentra el archivo "Cons-sem__.mgzn.16-05-23.xlsx".\n Si esta el archivo, verifique que el nombre sea correspondiente')
        sys.exit()

    name_counter = 0
    for worker_file in worker_files_list:
        # Adding summary information
        resume_sheet = files_names[name_counter] + '.r'
        resume_ws = None
        try:
            resume_ws = workbook[resume_sheet]
        except KeyError:
            # Exception: No hay una hoja con ese nombre en el modelo.
            messagebox.showerror(title='Error', message=
            'No se encuentra la hoja con nombre "'+ resume_sheet+'" en archivo Cons-sem__.mgzn.16-05-23.xlsx.\n')
            sys.exit()
        is_pivot = worker_file.get_name() == 'aaa'
        ProductsManage.set_products(resume_ws, worker_file, is_pivot=is_pivot)
        CoinInventoryManage.set_coin_inventory_source(resume_ws, worker_file, is_pivot=is_pivot)
        MerchandiseControlManage.set_merchandise_control_source(resume_ws, worker_file, is_pivot=is_pivot)
        #     Adding date in summary sheet
        resume_ws.cell(row=2, column=1, value=date)
        resume_ws.cell(row=3, column=1, value=week_number)

        # Adding balance information
        # Exception: No hay una hoja con ese nombre en el modelo
        balance_sheet = files_names[name_counter] + '.b'
        balance_ws = workbook[balance_sheet]

        BalanceManage.set_balance_source(balance_ws, worker_file, is_pivot=is_pivot)

        name_counter += 1
    current_Date = datetime.datetime.now()
    date = str(current_Date.day) + '-' + str(current_Date.month) \
           + '-' + str(current_Date.year)[2:] + '  [' + str(current_Date.hour) + \
           ';' + str(current_Date.minute) + ']'

    name_new_report = str(name_model)[:8]+str(week_number)+str(name_model)[10:]+date+'.xlsx'

    print('Guardado: ', name_new_report)
    workbook.save(name_new_report)
    # ChangeFormat.xlsx_to_xlsb(name_new_report)

    # delete all temporal files.
    # os.remove(temp_file_name)
    # os.remove(name_new_report)

    end = time.time()

    print('Tiempo de ejecucion: ', end - begin)
    messagebox.showinfo(title='Listo', message='Todo bien\n')
    sys.exit()

if __name__ == "__main__":
    consolidate()
